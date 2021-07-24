#Jone Huan Chong 201533109

from openpyxl import load_workbook, Workbook
class Card:

    def __init__(self, theName, theType, theHP, theMoves, isShiny): #initialize card
        self.name = theName
        self.type = theType
        self.HP = theHP
        self.moves = theMoves
        self.shiny = isShiny
        if self.shiny == 1: #represent shiny status of yes/no for sensible printing
            self.shinyStatus = "yes" 
        else:
            self.shinyStatus = "no"
        
    def __str__(self):
        return "Name:{self.name}, Type:{self.type}, HP:{self.HP}, Moves:{self.moves}, Shiny:{self.shinyStatus}".format(self=self)

class Deck:

    def __init__(self):
        self.deck = {} #dictionary used to initialize deck, enables objects to be referenced by their dictionary key name
    
    def inputFromFile(self, fileName):
        workbook = load_workbook(fileName).active        
        for row in workbook.iter_rows (min_row=2, max_row=7, values_only=True):
            movesInformation = {} #initialize dictionary of move names and damages, (to be nested later)
            movesInformation = {row[4] :  row[5],
                                row[6] :  row[7],
                                row[8] :  row[9],
                                row[10] : row[11],
                                row[12] : row[13]}
            for i in list(movesInformation): #deletes xlsx cells that are NoneType
                if i is None:
                    del movesInformation[i]

            nameInformation = row[0]
            typeInformation = row[1]
            hpInformation = row[2]
            shinyInformation = row[3]

            self.deck[nameInformation] = Card(nameInformation, typeInformation, hpInformation, movesInformation, shinyInformation) #initialize every card row into the Card class and add to dictionary
            
    def __str__(self):
        return "Total number of cards: " + str(len(self.deck)) +"\nTotal number of shiny cards: " + str(self.shinyCounter()) + "\nAverage damage of deck: " + str(self.getAverageDamage())

    def shinyCounter(self): #shinyCounter returns the number of shiny cards in the deck
        shinyCount = 0 
        for key in list((self.deck).keys()):
            if self.deck[key].shiny == 1:
                shinyCount += 1
        return shinyCount

    def addCard(self, theCard): #adds card object to the deck dictionary, with the the name of card as the key, and the object as the value
        self.deck[theCard.name] = theCard
    
    def rmCard(self, theCard): #removes the card from deck and prints the updated card list
        for key in list((self.deck).keys()):
            if self.deck[key].name == theCard.name:
                print("Card of name: " , self.deck[key].name, "has been deleted")
                del self.deck[key]

        print("Updated card list: ")       
        self.getCards()

    def getIndividualAverageDamage(self, referenceCard): #method to return the average damage of the specific card
        totalCardDamage = 0
        x = (self.deck)[referenceCard].moves
        for moveset in list(x):
            totalCardDamage += (x[moveset])
        individualAverageDamage = (totalCardDamage)/(len(x))
        return individualAverageDamage

    def getMostPowerful(self): #returns a list of the most powerful cards in the deck. a list is used as there might be a possbility of different cards having the same highest average damage
        cardAverage = {}
        mostPowerfulCards = []
        for key in list((self.deck).keys()):
            cardAverage[key] = self.getIndividualAverageDamage(key)
        for i in cardAverage:
            if cardAverage[i] == max(cardAverage.values()):
                mostPowerfulCards.append(i)
        return mostPowerfulCards

    def getAverageDamage(self): #returns the average damage of the deck
        totalDeckDamage = 0
        for key in list((self.deck).keys()):
            totalDeckDamage += self.getIndividualAverageDamage(key)
        deckAverageDamage = totalDeckDamage/(len(self.deck))
        return deckAverageDamage
                
    def viewAllCards(self): #prints all the cards in the deck in a sensible manner
        print("All cards and information of the deck: ")
        for key in list((self.deck).keys()):
            print(self.deck[key])        

    def viewAllShinyCards(self): #filters the deck for shiny cards and prints the names of only shiny cards
        print("Shiny Cards in the deck:")
        for key in list((self.deck).keys()):
            if self.deck[key].shiny == 1:
                print(self.deck[key])
            
    def viewAllByType(self, theType): #prints all cards of the deck that has the same type specfied
        print("Cards of type ",theType,": ")
        for key in list((self.deck).keys()):
            if self.deck[key].type == theType:
                print(self.deck[key])
                
    def getCards(self): #prints the name of cards in the deck
        print("List of card names in the deck:")
        for key in list((self.deck).keys()):
            print(self.deck[key].name)
        
    def saveToFile(self, fileName):
        wb = Workbook()
        ws = wb.active

        ws.cell(row=1,column=1). value= "Name" #define the labels in the first row of the xlsx file
        ws.cell(row=1,column=2). value= "Type"
        ws.cell(row=1,column=3). value= "HP"
        ws.cell(row=1,column=4). value= "Shiny"
        ws.cell(row=1,column=5). value= "Move Name 1"
        ws.cell(row=1,column=6). value= "Damage 1"
        ws.cell(row=1,column=7). value= "Move Name 2"
        ws.cell(row=1,column=8). value= "Damage 2"        
        ws.cell(row=1,column=9). value= "Move Name 3"
        ws.cell(row=1,column=10). value= "Damage 3"
        ws.cell(row=1,column=11). value= "Move Name 4"
        ws.cell(row=1,column=12). value= "Damage 4"
        ws.cell(row=1,column=13). value= "Move Name 5"
        ws.cell(row=1,column=14). value= "Damage 5"
     
        rowCount = 1  #formatting of rows     
        for key in list((self.deck).keys()): #loop through every object and define the cell values
            ws.cell(row=rowCount+1,column=1). value= str(self.deck[key].name)
            ws.cell(row=rowCount+1,column=2). value= str(self.deck[key].type)
            ws.cell(row=rowCount+1,column=3). value=str( self.deck[key].HP)    
            ws.cell(row=rowCount+1,column=4). value= str(self.deck[key].shiny)
            
            moveColCount = 5 #formatting of move name column
            dmgColCount = 6 #formatting of move damage column
            for move in list(self.deck[key].moves): #loop through objects's moves
                ws.cell(row=rowCount+1,column= moveColCount).value = move
                ws.cell(row=rowCount+1,column= dmgColCount).value = (self.deck[key].moves)[move]
                moveColCount += 2
                dmgColCount +=2
            rowCount += 1
        wb.save(fileName+".xlsx") #saves file into same working folder

def main():
    # d1 = Deck()
    # d1.inputFromFile("sampleDeck.xlsx")
    # print(d1.deck)
    # c1 = Card("A", "Magi", 1000, {"m1" : 6, "m2" : 66, "m3" : 666}, 0)
    # d1.addCard(c1)
    # c2 = Card("B", "Magi", 1000, {"m1" : 7, "m2" : 66, "m3" : 166}, 0)
    # d1.addCard(c2)
    # d1.rmCard(c1)
    # print(d1.deck["B"].HP)
    # print(c2)
    # print(d1.deck["Hoodoos"])
    # d1.viewAllShinyCards()
    # d1.viewAllByType("Magi")
    # d1.getIndividualAverageDamage("Hoodoos")
    # d1.getAverageDamage()
    # d1.getMostPowerful()
    # print(d1)
    # d1.viewAllCards()
    # d1.viewAllByType("Magi")
    # d1.saveToFile("NewDeck")


if __name__ == "__main__":
      main()